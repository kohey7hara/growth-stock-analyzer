"""
generate_report.py - Excel分析レポート自動生成モジュール

分析結果CSVからプロフェッショナルなExcelレポートを生成する。
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

BASE_DIR = Path(__file__).parent

# スタイル定義
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(name="Arial", bold=True, size=12, color="333333")
DATA_FONT = Font(name="Arial", size=10)
RED_FONT = Font(name="Arial", size=10, color="FF0000")
GREEN_FONT = Font(name="Arial", size=10, color="008000")
STRONG_BUY_FILL = PatternFill("solid", fgColor="C6EFCE")
BUY_FILL = PatternFill("solid", fgColor="E2EFDA")
WATCH_FILL = PatternFill("solid", fgColor="FFF2CC")
HOLD_FILL = PatternFill("solid", fgColor="FCE4EC")
ZEBRA_FILL = PatternFill("solid", fgColor="F2F7FB")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)


def signal_fill(sig):
    if "強い買い" in str(sig): return STRONG_BUY_FILL
    elif "買い" in str(sig): return BUY_FILL
    elif "要注目" in str(sig): return WATCH_FILL
    return HOLD_FILL


def write_header_row(ws, row, headers, widths=None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 32
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def write_data_row(ws, row_num, values, idx, signal_col=0, change_col=None, wrap_cols=None):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=v)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(col in (wrap_cols or [])))
        if idx % 2 == 1:
            cell.fill = ZEBRA_FILL

    if signal_col:
        ws.cell(row=row_num, column=signal_col).fill = signal_fill(values[signal_col - 1])
        ws.cell(row=row_num, column=signal_col).font = Font(name="Arial", bold=True, size=10)

    if change_col:
        cell = ws.cell(row=row_num, column=change_col)
        try:
            val = float(cell.value) if cell.value else 0
            cell.font = GREEN_FONT if val > 0 else RED_FONT if val < 0 else DATA_FONT
        except (ValueError, TypeError):
            pass

    ws.row_dimensions[row_num].height = 50


def _safe_val(r, key, default=""):
    """DataFrameの値を安全に取得。nanの場合はdefaultを返す"""
    val = r.get(key)
    if val is None:
        return default
    try:
        if pd.isna(val):
            return default
    except (TypeError, ValueError):
        pass
    return val


def generate_excel_report(analysis_df, output_path=None, macro_df=None, portfolio_data=None):
    """分析結果DataFrameからExcelレポートを生成"""
    if output_path is None:
        output_path = BASE_DIR / "output" / "growth_stock_report.xlsx"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    now = datetime.now()

    # ============================================================
    # Sheet 1: 総合ランキング (改善3: analysis_comment使用)
    # ============================================================
    ws = wb.active
    ws.title = "総合ランキング"

    ws.merge_cells("A1:N1")
    ws["A1"] = "グロース株 総合分析レポート"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:N2")
    ws["A2"] = f"分析日時: {now.strftime('%Y年%m月%d日 %H:%M')}    ※本レポートは情報提供であり投資助言ではありません"
    ws["A2"].font = Font(name="Arial", size=9, color="666666")

    headers = [
        "順位", "シグナル", "ティッカー", "企業名", "市場", "セクター",
        "株価", "前日比(%)", "総合スコア",
        "テクニカル", "ファンダ", "センチメント", "モメンタム",
        "主要分析コメント"
    ]
    widths = [6, 16, 10, 20, 6, 16, 14, 10, 10, 10, 10, 12, 10, 70]
    write_header_row(ws, 4, headers, widths)

    for i, (_, r) in enumerate(analysis_df.iterrows()):
        row_num = 5 + i
        # 改善3: analysis_comment があればそれを使用
        main_comment = _safe_val(r, "analysis_comment", "")
        if not main_comment:
            # フォールバック: 従来のreasons結合
            comments = []
            for key in ["tech_reasons", "fund_reasons", "sent_reasons", "mom_reasons"]:
                val = r.get(key, "")
                if pd.notna(val) and val:
                    comments.append(str(val))
            main_comment = " | ".join(comments[:3])

        price_display = r.get("price", "")
        if r.get("market") == "JP":
            price_display = f"¥{int(price_display):,}" if pd.notna(price_display) else ""
        else:
            price_display = f"${price_display:,.2f}" if pd.notna(price_display) else ""

        vals = [
            i + 1,
            r.get("signal", ""),
            r.get("ticker", ""),
            r.get("name", ""),
            r.get("market", ""),
            r.get("sector", ""),
            price_display,
            r.get("change_pct", ""),
            r.get("total_score", ""),
            r.get("tech_score", ""),
            r.get("fund_score", ""),
            r.get("sent_score", ""),
            r.get("mom_score", ""),
            main_comment,
        ]
        write_data_row(ws, row_num, vals, i, signal_col=2, change_col=8, wrap_cols=[14])

    # ============================================================
    # Sheet 2: テクニカル詳細 (改善4: 出来高実数値追加)
    # ============================================================
    ws2 = wb.create_sheet("テクニカル詳細")
    ws2.merge_cells("A1:P1")
    ws2["A1"] = "テクニカル指標一覧"
    ws2["A1"].font = TITLE_FONT
    ws2.row_dimensions[1].height = 35

    tech_headers = [
        "ティッカー", "企業名", "株価", "RSI(14)", "MACD", "MACDヒストグラム",
        "BB位置(%)", "SMA20", "SMA50", "SMA200", "クロスシグナル",
        "当日出来高", "20日平均出来高", "出来高比率", "出来高状態", "過去5日出来高"
    ]
    tech_widths = [10, 18, 14, 10, 12, 14, 10, 14, 14, 14, 16, 16, 16, 10, 14, 30]
    write_header_row(ws2, 3, tech_headers, tech_widths)

    for i, (_, r) in enumerate(analysis_df.iterrows()):
        row_num = 4 + i

        # 出来高比率の表示と状態
        vol_ratio = _safe_val(r, "vol_ratio", "")
        vol_status = ""
        if vol_ratio and vol_ratio != "":
            try:
                vr = float(vol_ratio)
                if vr >= 2.0:
                    vol_status = "出来高急増！"
                elif vr >= 1.5:
                    vol_status = "増加"
                elif vr >= 0.8:
                    vol_status = "通常"
                else:
                    vol_status = "減少"
            except (ValueError, TypeError):
                pass

        # 出来高を見やすくフォーマット
        volume = _safe_val(r, "volume", "")
        vol_avg = _safe_val(r, "vol_avg_20d", "")
        if volume and volume != "":
            try:
                volume = f"{int(volume):,}"
            except (ValueError, TypeError):
                pass
        if vol_avg and vol_avg != "":
            try:
                vol_avg = f"{int(vol_avg):,}"
            except (ValueError, TypeError):
                pass

        # 過去5日出来高
        vol_5d = _safe_val(r, "vol_5d", "")
        if vol_5d and vol_5d != "":
            try:
                vol_5d_list = [f"{int(v):,}" for v in str(vol_5d).split(",") if v.strip()]
                vol_5d = " → ".join(vol_5d_list)
            except (ValueError, TypeError):
                pass

        vals = [
            r.get("ticker"),
            r.get("name"),
            r.get("price"),
            r.get("rsi_14"),
            r.get("macd"),
            r.get("macd_histogram"),
            r.get("bb_position"),
            r.get("sma_20"),
            r.get("sma_50"),
            r.get("sma_200"),
            r.get("cross_signal"),
            volume,
            vol_avg,
            vol_ratio,
            vol_status,
            vol_5d,
        ]
        write_data_row(ws2, row_num, vals, i, wrap_cols=[16])

        # 改善4: 出来高比率2.0以上のセルを黄色ハイライト
        try:
            vr_val = float(_safe_val(r, "vol_ratio", 0))
            if vr_val >= 2.0:
                for col in [12, 13, 14, 15]:  # 出来高関連の列
                    cell = ws2.cell(row=row_num, column=col)
                    cell.fill = YELLOW_FILL
                    cell.font = Font(name="Arial", bold=True, size=10, color="FF0000")
        except (ValueError, TypeError):
            pass

    # ============================================================
    # Sheet 3: 株価推移 (改善1: 新規追加)
    # ============================================================
    ws_hist = wb.create_sheet("株価推移")
    ws_hist.merge_cells("A1:M1")
    ws_hist["A1"] = "過去5年間の株価推移"
    ws_hist["A1"].font = TITLE_FONT
    ws_hist.row_dimensions[1].height = 35

    hist_headers = [
        "ティッカー", "企業名", "現在株価",
        "5年前株価", "3年前株価", "1年前株価",
        "5年騰落率(%)", "3年騰落率(%)", "1年騰落率(%)",
        "1年最高値", "最高値日", "1年最安値", "最安値日"
    ]
    hist_widths = [10, 18, 14, 14, 14, 14, 12, 12, 12, 14, 12, 14, 12]
    write_header_row(ws_hist, 3, hist_headers, hist_widths)

    for i, (_, r) in enumerate(analysis_df.iterrows()):
        row_num = 4 + i

        # 株価フォーマット
        def fmt_price(val, market):
            v = _safe_val(r, val, "")
            if v == "" or v is None:
                return ""
            try:
                if market == "JP":
                    return f"¥{int(float(v)):,}"
                else:
                    return f"${float(v):,.2f}"
            except (ValueError, TypeError):
                return ""

        market = r.get("market", "US")
        current_price = fmt_price("price", market)

        # 騰落率に色付け用の値
        def fmt_change(key):
            v = _safe_val(r, key, "")
            if v == "" or v is None:
                return ""
            try:
                return round(float(v), 1)
            except (ValueError, TypeError):
                return ""

        vals = [
            r.get("ticker"),
            r.get("name"),
            current_price,
            fmt_price("price_5y_ago", market),
            fmt_price("price_3y_ago", market),
            fmt_price("price_1y_ago", market),
            fmt_change("change_5y_pct"),
            fmt_change("change_3y_pct"),
            fmt_change("change_1y_pct"),
            fmt_price("high_1y", market),
            _safe_val(r, "high_1y_date", ""),
            fmt_price("low_1y", market),
            _safe_val(r, "low_1y_date", ""),
        ]
        write_data_row(ws_hist, row_num, vals, i)

        # 騰落率に色をつける
        for col_idx in [7, 8, 9]:  # 5年/3年/1年騰落率
            cell = ws_hist.cell(row=row_num, column=col_idx)
            try:
                val = float(cell.value) if cell.value else 0
                cell.font = GREEN_FONT if val > 0 else RED_FONT if val < 0 else DATA_FONT
            except (ValueError, TypeError):
                pass

    # ============================================================
    # Sheet 4: センチメント詳細
    # ============================================================
    ws3 = wb.create_sheet("センチメント詳細")
    ws3.merge_cells("A1:J1")
    ws3["A1"] = "SNS・話題性データ"
    ws3["A1"].font = TITLE_FONT
    ws3.row_dimensions[1].height = 35

    sent_headers = [
        "ティッカー", "企業名", "複合センチメント",
        "Trends比率", "Trends急上昇",
        "Redditメンション", "Redditセンチメント",
        "Twitterメンション", "Twitterセンチメント",
        "センチメント分析"
    ]
    sent_widths = [10, 18, 14, 10, 12, 14, 14, 14, 14, 40]
    write_header_row(ws3, 3, sent_headers, sent_widths)

    for i, (_, r) in enumerate(analysis_df.iterrows()):
        row_num = 4 + i
        gt_trending = r.get("gtrends_is_trending")
        if pd.isna(gt_trending) if isinstance(gt_trending, float) else not gt_trending:
            trending_display = "-"
        else:
            trending_display = "急上昇!" if gt_trending else "-"

        vals = [
            r.get("ticker"),
            r.get("name"),
            r.get("combined_sentiment"),
            r.get("gtrends_trend_ratio"),
            trending_display,
            r.get("reddit_mentions"),
            r.get("reddit_sentiment"),
            r.get("twitter_mentions"),
            r.get("twitter_sentiment"),
            r.get("sent_reasons", ""),
        ]
        write_data_row(ws3, row_num, vals, i, wrap_cols=[10])

    # ============================================================
    # Sheet 5: スコアリング方法
    # ============================================================
    ws4 = wb.create_sheet("スコアリング方法")
    ws4["A1"] = "スコアリング方法と配分"
    ws4["A1"].font = TITLE_FONT
    ws4.row_dimensions[1].height = 35

    ws4["A3"] = "■ スコア配分"
    ws4["A3"].font = SUBTITLE_FONT

    alloc = [
        ["カテゴリ", "配分", "最高点", "主な評価項目"],
        ["テクニカル分析", "35%", "100点", "RSI / MACD / ボリンジャーバンド / ゴールデンクロス"],
        ["ファンダメンタルズ", "30%", "100点", "52週レンジ位置 / PER・PEG / 売上成長率 / 目標株価"],
        ["SNSセンチメント", "20%", "100点", "Google Trends / Reddit / X(Twitter) メンション・センチメント"],
        ["モメンタム", "15%", "100点", "出来高比率(20日平均比) / 移動平均線トレンド"],
    ]
    for i, row_data in enumerate(alloc):
        r = 4 + i
        for col, v in enumerate(row_data, 1):
            cell = ws4.cell(row=r, column=col, value=v)
            cell.border = THIN_BORDER
            cell.font = HEADER_FONT if i == 0 else DATA_FONT
            if i == 0: cell.fill = HEADER_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws4["A10"] = "■ シグナル判定"
    ws4["A10"].font = SUBTITLE_FONT
    signals = [
        ["70点以上", "★★★ 強い買い", "複数指標で強いシグナル。積極的にエントリーを検討"],
        ["55-69点", "★★ 買い", "全体的に良好。タイミングを見て買い"],
        ["40-54点", "★ 要注目", "一部に魅力あり。押し目を待つ"],
        ["39点以下", "△ 様子見", "現時点では見送り推奨"],
    ]
    for i, sd in enumerate(signals):
        r = 11 + i
        for col, v in enumerate(sd, 1):
            cell = ws4.cell(row=r, column=col, value=v)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    for i, w in enumerate([18, 12, 22, 55], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ============================================================
    # Sheet 6: マーケット環境 (改善2)
    # ============================================================
    if macro_df is not None and not macro_df.empty:
        ws_macro = wb.create_sheet("マーケット環境")
        ws_macro.merge_cells("A1:H1")
        ws_macro["A1"] = "マーケット環境サマリー"
        ws_macro["A1"].font = TITLE_FONT
        ws_macro.row_dimensions[1].height = 35

        ws_macro["A2"] = f"取得日時: {now.strftime('%Y年%m月%d日 %H:%M')}"
        ws_macro["A2"].font = Font(name="Arial", size=9, color="666666")

        macro_headers = [
            "指標", "カテゴリ", "現在値", "前日比(%)", "1ヶ月変動(%)", "3ヶ月変動(%)"
        ]
        macro_widths = [22, 14, 16, 12, 14, 14]
        write_header_row(ws_macro, 4, macro_headers, macro_widths)

        for i, (_, r) in enumerate(macro_df.iterrows()):
            row_num = 5 + i
            vals = [
                r.get("name", ""),
                r.get("category", ""),
                r.get("current_value", ""),
                r.get("change_1d_pct", ""),
                r.get("change_1m_pct", ""),
                r.get("change_3m_pct", ""),
            ]
            write_data_row(ws_macro, row_num, vals, i)

            # 変動率に色をつける
            for col_idx in [4, 5, 6]:
                cell = ws_macro.cell(row=row_num, column=col_idx)
                try:
                    val = float(cell.value) if cell.value else 0
                    cell.font = GREEN_FONT if val > 0 else RED_FONT if val < 0 else DATA_FONT
                except (ValueError, TypeError):
                    pass

        # 市場環境の解釈コメント
        interp_row = 5 + len(macro_df) + 2
        ws_macro.cell(row=interp_row, column=1, value="■ 市場環境の解釈").font = SUBTITLE_FONT

        try:
            from fetch_macro import interpret_macro
            interpretations = interpret_macro(macro_df)
            for j, interp in enumerate(interpretations):
                r = interp_row + 1 + j
                ws_macro.cell(row=r, column=1, value=interp["indicator"]).font = Font(name="Arial", bold=True, size=10)
                ws_macro.cell(row=r, column=2, value=interp["value"]).font = DATA_FONT
                ws_macro.cell(row=r, column=3, value=interp["interpretation"]).font = Font(name="Arial", bold=True, size=10, color="1F4E79")
                ws_macro.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
                ws_macro.cell(row=r, column=4, value=interp["impact"]).font = DATA_FONT
                ws_macro.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
        except Exception:
            pass

    # ============================================================
    # Sheet 7: ポートフォリオ (改善4)
    # ============================================================
    if portfolio_data and portfolio_data[0] is not None:
        port_df, totals = portfolio_data
        ws_port = wb.create_sheet("ポートフォリオ")
        ws_port.merge_cells("A1:N1")
        ws_port["A1"] = "保有銘柄一覧・損益状況"
        ws_port["A1"].font = TITLE_FONT
        ws_port.row_dimensions[1].height = 35

        usdjpy_str = f"  USDJPY: {totals.get('usdjpy_rate', 'N/A')}" if totals.get('usdjpy_rate') else ""
        ws_port["A2"] = f"分析日時: {now.strftime('%Y年%m月%d日 %H:%M')}{usdjpy_str}"
        ws_port["A2"].font = Font(name="Arial", size=9, color="666666")

        port_headers = [
            "ティッカー", "企業名", "購入日", "購入単価", "数量",
            "現在株価", "評価損益(%)", "損益額", "保有日数",
            "年率リターン(%)", "評価額(円)", "1年後予想株価", "メモ"
        ]
        port_widths = [10, 18, 12, 14, 8, 14, 12, 14, 10, 12, 16, 14, 25]
        write_header_row(ws_port, 4, port_headers, port_widths)

        for i, (_, r) in enumerate(port_df.iterrows()):
            row_num = 5 + i
            vals = [
                r.get("ticker", ""),
                r.get("name", ""),
                r.get("buy_date", ""),
                r.get("buy_price", ""),
                r.get("quantity", ""),
                r.get("current_price", ""),
                r.get("unrealized_pnl_pct", ""),
                r.get("unrealized_pnl", ""),
                r.get("holding_days", ""),
                r.get("annualized_return", ""),
                r.get("current_value_jpy", ""),
                r.get("projected_price_1y", ""),
                r.get("memo", ""),
            ]
            write_data_row(ws_port, row_num, vals, i, wrap_cols=[13])

            # 損益に色をつける
            for col_idx in [7, 8, 10]:
                cell = ws_port.cell(row=row_num, column=col_idx)
                try:
                    val = float(cell.value) if cell.value else 0
                    cell.font = GREEN_FONT if val > 0 else RED_FONT if val < 0 else DATA_FONT
                except (ValueError, TypeError):
                    pass

        # 合計行
        total_row = 5 + len(port_df) + 1
        ws_port.cell(row=total_row, column=1, value="合計").font = Font(name="Arial", bold=True, size=11)
        ws_port.cell(row=total_row, column=8, value=totals.get("total_pnl_jpy", "")).font = Font(
            name="Arial", bold=True, size=11,
            color="008000" if totals.get("total_pnl_jpy", 0) > 0 else "FF0000"
        )
        ws_port.cell(row=total_row, column=11, value=totals.get("total_current_jpy", "")).font = Font(name="Arial", bold=True, size=11)

        pnl_pct = totals.get("total_pnl_pct", 0)
        ws_port.cell(row=total_row, column=7, value=f"{pnl_pct:.1f}%" if pnl_pct else "").font = Font(
            name="Arial", bold=True, size=11,
            color="008000" if pnl_pct and pnl_pct > 0 else "FF0000"
        )

        # 免責事項
        disc_row = total_row + 2
        ws_port.merge_cells(start_row=disc_row, start_column=1, end_row=disc_row, end_column=13)
        ws_port.cell(row=disc_row, column=1,
                     value="※過去のリターンは将来を保証しません。1年後予想株価は過去3年間の成長率に基づく参考値です。").font = Font(
            name="Arial", size=9, color="999999", italic=True
        )

    wb.save(output_path)
    print(f"レポート保存: {output_path}")
    return output_path


if __name__ == "__main__":
    data_dir = BASE_DIR / "data"
    analysis_path = data_dir / "latest_analysis.csv"

    if not analysis_path.exists():
        print("分析結果がありません。先に analyzer.py を実行してください")
        exit(1)

    df = pd.read_csv(analysis_path)
    output = generate_excel_report(df)
    print(f"\nExcelレポート生成完了: {output}")
